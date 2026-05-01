from openpyxl import Workbook, load_workbook
import os

ARQUIVO_EXCEL = "contatos.xlsx"

def salvar_em_excel(dados):
    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contatos"

        ws.append(["Nome", "Email", "Telefone", "Empresa", "CNPJ", "Mensagem"])
    else:
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active

    ws.append([
        dados.nome,
        dados.email,
        dados.telefone,
        dados.empresa,
        dados.cnpj,
        dados.mensagem
    ])

    wb.save(ARQUIVO_EXCEL)